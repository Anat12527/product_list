import string
from datetime import date, datetime, timedelta
import re
import openpyxl
import pandas as pd
from e4_form import Logging
from flask import Flask, render_template, request, redirect, url_for, flash, session
from flask_bootstrap import Bootstrap
from flask_login import LoginManager, UserMixin, login_user, login_required, current_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from openpyxl.styles import Font
from sqlalchemy import create_engine,desc


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:newrootpassword@localhost/home_products_management'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'mysecretkey4'
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(minutes=10)
login_manager = LoginManager(app)
login_manager.login_view = 'log' #the user returns to this view function.
login_manager.login_message = 'You must login in order to access the app '
db = SQLAlchemy(app)
app.debug = True
Bootstrap(app)



###  Models  ###
################

class Departments(db.Model):
    """
       This is a class for product departments.

       Attributes:
           department_id (int): The department number.
           department_name(string): The department name.
    """
    __tablename__ = 'departments'
    department_id = db.Column('Id_Department', db.Integer, primary_key=True)
    department_name = db.Column('Department_Name', db.String(50), nullable=False)

    def __init__(self, department_id, department_name):
        self.department_id = department_id
        self.department_name = department_name

    """ 
        The constructor for Departments class. 

        Parameters: 
               department_id (int): The department number.
               department_name(string): The department name.   
    """


class Products(db.Model):
    """
           This is a class for products and their details.

           Attributes:

               product_name(string): The product name.
               department_id (int) : The department number.
               product_amount (int) : The product amount.
               product_notes (string) : The notes added for purchasing.
               date_buy (date) :The date a product was added to the list.(update is possible).


    """

    __tablename__ = 'products'
    product_id = db.Column('Id_Product', db.Integer, primary_key=True)
    product_name = db.Column('Product_Name', db.String(50), nullable=False)
    department_id = db.Column('Id_Department', db.Integer, db.ForeignKey('departments.Id_Department'))
    product_amount = db.Column('Product_amount', db.Integer, nullable=False)
    product_notes = db.Column('Product_Notes', db.String(50), nullable=False)
    date_buy = db.Column('Date_Buy', db.DATE, default=date.today(), nullable=False)
    deps = db.relationship('Departments', backref='prod')#could have benn used in the index_product to directly approach the departments name: product_details.deps.departmet_name

    """ 
            The constructor for Products class. 

            Parameters: 
               
               product_name(string): The product name.
               department_id (int) : The department number.
               product_amount (int) : The product amount.
               product_notes (string) : The notes added for purchasing.
               date_buy (date) :The date a product was added to the list.(update is possible).
        """

    def __init__(self, product_name, department_id, product_amount, product_notes, date_buy):
        self.department_id = department_id
        self.product_name = product_name
        self.product_amount = product_amount
        self.product_notes = product_notes
        self.date_buy = date_buy


class Recipes(db.Model):
     __tablename__ = 'recipes'
     recipe_id = db.Column('Id_Recipe', db.Integer, primary_key=True)
     recipe_name = db.Column('Recipe_Name', db.String(50), nullable=False)

     def __init__(self, recipe_name, recipe_id):
       self.recipe_id = recipe_id
       self.recipe_name = recipe_name


class ProdForRecipe(db.Model):
     __tablename__ = 'prodforrecipe'
     product_id = db.Column('Id_Product', db.Integer, primary_key=True)
     product_name = db.Column('Product_Name', db.String(50), nullable=False)
     department_id_rec = db.Column('Id_Department_rec', db.Integer, db.ForeignKey('departments.Id_Department'))
     product_amount = db.Column('Product_amount', db.Integer, nullable=False)
     product_notes = db.Column('Product_Notes', db.String(50), nullable=False)
     recipe_id = db.Column('Id_Recipe', db.Integer, db.ForeignKey('recipes.Id_Recipe'))
     deps2 = db.relationship('Departments', backref='prod2')
     rec_name = db.relationship('Recipes', backref='prod_rec')

     def __init__(self, product_name, department_id_rec, product_amount, product_notes,recipe_id):
         self. department_id_rec =  department_id_rec
         self.product_name = product_name
         self.product_amount = product_amount
         self.product_notes = product_notes
         self.recipe_id = recipe_id



class Users(db.Model, UserMixin):
    """
               This is a class for users names and passwords.

               Attributes:

                   id_user(int): The user number.
                   user_name (string) : The user name.
                   password (string) : The user password.

        """

    __tablename__ = 'users'
    id_user = db.Column('Id_User', db.Integer, primary_key=True)
    user_name = db.Column('User_Name', db.String(50), nullable=False)
    password = db.Column('Password', db.String(50), nullable=False)

    def __init__(self, id_user, user_name, password):
        self.id_user = id_user
        self.user_name = user_name
        self.password = password

    """ 
               The constructor for Users class. 

               Parameters: 

                  id_user(int): The user number.
                  user_name (string) : The user name.
                  password (string) : The user password.

    """

    def get_id(self):
        """Returns id_user parameter,used because class usermixin refers to id.(returns the id of the table)."""
        return (self.id_user)


class QueriesMyDb():
    """
            This is a class for querying by filter name, and filter products by department  names .

            Attributes:

                      choosen_table(string): The main able referred in the query.
                      parameter_f (string) : The parameter which is being filtered .
                      colum_name (string) : The column name from which needs to be filtered .

           """

    def __init__(self, choosen_table, parameter_f, colum_name):
        self.choosen_table = choosen_table
        self.parameter_f = parameter_f
        self.colum_name = colum_name

        """ 
            The constructor for QueriesMyDb class. 

            Parameters: 

                      choosen_table(string): The main able referred in the query.
                      parameter_f (string) : The parameter which is being filtered .
                      colum_name (string) : The column name from which needs to be filtered .

           """

    def query_filter_name(self):
        """Returns an object named one_record_details, which is a record from the database."""
        one_record_details = db.session.query(self.choosen_table).filter(self.colum_name == self.parameter_f).first()
        return one_record_details

    def filter_products_by_dept_name(self):
        """Returns record/records as object/s named all_products_data_by_dept of products filtered by the department they where attached to."""

        all_products_data_by_dept = db.session.query(self.choosen_table).join(Departments,
                                                                              Departments.department_id == Products.department_id).filter(
            self.colum_name == self.parameter_f).all()
        return all_products_data_by_dept


#### view functions  ####
######################################
@login_manager.user_loader
def get_user_id(id_user):
    """Returns an the user id as integer.receives id_user  ."""
    return Users.query.get(int(id_user))


@app.route('/log', methods=['GET', 'POST'])
def log():
    """
       verifies user name and password according to the database.


       Returns:
       redirect: to col_names_dict_names_to_session, which builds  the list of products if password and user name are correct.
       render_template : returns back to log in page with  matching error message in the terms user mame or password were incorrect.

       """
    error_m = None
    form = Logging()

    if form.validate_on_submit():
        user_name_entered = form.name.data
        user_name_password = form.password.data
        user = QueriesMyDb(Users, user_name_entered, Users.user_name)
        user1 = user.query_filter_name()
        if user1:
            login_user(user1, remember=True)#remember is true in order to remember the user after the session expires.
            if user_name_password == user1.password:
                return redirect(url_for('col_names_dict_names_to_session'))
            else:
                error_m = ["The password is incorrect", ""]
                return render_template('log_in.html', form=form, error_m=error_m)
        else:
            error_m = ["", "The user name is incorrect"]
            return render_template('log_in.html', form=form, error_m=error_m)
    else:
        return render_template('log_in.html', form=form, error_m=error_m)


@app.route('/logout')
@login_required
def logout():
    """
      logs out the user.

      Returns:
       redirect: to the page of lof in  ."""
    logout_user()
    return redirect(url_for('log'))


@app.route('/', methods=['POST', 'GET'])
@login_required
def col_names_dict_names_to_session():
    """
          inserts department names as values and department numbers as keys into dictionary.
          inserts Products column names into a list.
          inserts Users column names into  a list.
          all inserts mentioned above are stored as sessions.

          Returns:
          redirect: to the list of products if password and user name are correct.
          """

    dict_departments = {r.department_id: r.department_name for r in Departments.query.all()}
    session['dict_departments'] = dict_departments #insering to a temporary variable.
    list_column_names = Products.__table__.columns.keys()
    session['list_column_names'] = list_column_names
    list_colum_names_users = Users.__table__.columns.keys()
    session['list_colum_names_users'] = list_colum_names_users
    list_column_names_prod_recipes = ProdForRecipe.__table__.columns.keys()
    session['list_column_names_prod_recipes'] = list_column_names_prod_recipes
    dict_recipes = {rec.recipe_id:rec.recipe_name for rec in Recipes.query.all()}
    session['dict_recipes'] = dict_recipes
    list_column_names_recipes = Recipes.__table__.columns.keys()
    session['list_column_names_recipes'] = list_column_names_recipes
    return redirect(url_for('show_products'))


@app.context_processor
def insert_date_today():
    """
              inserts date into html templates.

              Returns:
               dictionary  name: today_d .
     """
    now = datetime.now()
    today_string = now.strftime('%d-%m-%Y')
    return {'today_d': today_string}#a const variable that is used for the export view.


@app.context_processor  # returns the sheets names.puts names in a dictionary and the outcome dictionary in another dict.
def insert_name_columns():
    """
              tries to open the excel file. If succeed, it inserts list of sheet names in html templates.
              Returns:
              dictionary named dict_sheet_names .
              """
    dict_sheet_names = dict()
    try:
        recievef = openpyxl.load_workbook('product_lists.xlsx')
    except FileNotFoundError:
        flash(
            "The file 'product_lists.xlsx' was not found. please create it.The export and import utile will be out of order. ")

    else:
        g_sheets_f = recievef.sheetnames
        dict_sheet_names = dict()
        for index, sheet_name in enumerate(g_sheets_f):
            dict_sheet_names[index] = sheet_name


    finally:
        return {'dict_sheet_names': dict_sheet_names}


@app.route('/index_products/get_data', methods=['POST', 'GET'])
@login_required
def inserts_products_to_dict(num_department, dict_departments, dict_products_by_department):
    """
    
        queries all products by their number department and insert them to a dict.each key is a department. the values are the products objects.
        Parameters:
        num_department (int): the number of the department.
        dict_departments(dict) : a dict with department name as value.
        dict_products_by_department(dict) : a dict that will contain department name and products objects as values.
    
    
        Returns:
        int: Description of return value
  
    """

    Qall = QueriesMyDb(Products, dict_departments[num_department],
                       Departments.department_name)  # queries department  and joins number department in products.
    all_products_data_by_dept = Qall.filter_products_by_dept_name()
    if all_products_data_by_dept:  # if data found in certain department.
        dict_products_by_department[dict_departments[num_department]] = all_products_data_by_dept
    return dict_products_by_department


@app.route('/index_products', methods=['POST', 'GET'])  # shows the list of products
@login_required
def show_products():
    """ 

        shows all products by their number department and insert them to a dict.
        
        Returns: 
        return template : returns  the template of the product list. 

        """

    find_name = None
    find_dept = None
    dict_products_by_department = dict()
    dict_departments = session['dict_departments']
    list_column_names = session['list_column_names']
    for num_department in dict_departments:  # checks for each  number department if there are products in the list  that belong to it.
        dict_products_by_department = inserts_products_to_dict(num_department, dict_departments,
                                                               dict_products_by_department)
    return render_template('index_products.html', list_column_names=list_column_names,
                           dict_products_by_department=dict_products_by_department, dict_departments=dict_departments,
                           find_name=find_name, find_dept=find_dept, logged_user=current_user.user_name)


@app.route('/index_users', methods=['POST', 'GET'])
@login_required
def show_all_users():
    """ 
        queries all users. And showing them as a table.
        
        Returns: 
        render_template: the table of users. 

        """
    list_colum_names_users = session['list_colum_names_users']
    all_users = Users.query.all()
    return render_template('index_users.html', all_users=all_users, list_colum_names_users=list_colum_names_users,
                           logged_user=current_user.user_name)


@app.route('/add_users', methods=['POST', 'GET'])
@login_required
def add_users():
    """ 

        Adds users,if username doesn't exist and is not numeric.

        Returns: 
        redirect: returns to the users page.

        """
    if request.method == 'POST':
        user_name_entered = request.form.get('user_name')
        user_password_entered = request.form.get('password')
        if user_name_entered.isalpha():
            user = QueriesMyDb(Users, user_name_entered, Users.user_name)
            user_query = user.query_filter_name()
            if user_query:
                flash("The user " + user_name_entered + " exists and can not be added", "warning")
            else:
                new_user = Users(None, user_name_entered, user_password_entered)
                db.session.add(new_user)
                db.session.commit()
        else:
            flash("The user name cannot consist numbers", "warning")
    return redirect(url_for('show_all_users'))


@app.route('/del_user/<int:user_id>', methods=['POST', 'GET'])
@login_required
def del_user(user_id):
    """

        deletes user if the user to be deleted is not logged in.
        Parameters:
        user_id (int): the user id.

        Returns:
        redirect: the html page, which shows all users.

        """

    if request.method == 'GET':
        if user_id == current_user.id_user:
            flash("The user " + current_user.user_name + " is logged in and can not be deleted", "warning")
        else:
            u_to_delete = Users.query.get(user_id)
            db.session.delete(u_to_delete)
            db.session.commit()
    return redirect(url_for('show_all_users'))


@app.route('/add_product/get_details', methods=['POST', 'GET'])
def get_details_product():
    """
            gets the details of  anew product.

            Returns:
            product_name_entered(string): The name of the product.
            department_id_entered(int): The number of the department.
            amount_entered (int): the amount of the product.
            notes_entered(string): the notes for each product.

            """

    product_name_entered = request.form.get('product_name')
    department_name_entered = request.form.get('department_select')  # user selects department name
    dict_departments = session['dict_departments']
    for key_id in dict_departments:  # translating department name chosen to department id.
        if department_name_entered == dict_departments[
            key_id]:  # if finds the name of department in dict.get its number.
            department_id_entered = key_id
    amount_entered = request.form.get('amount')
    notes_entered = request.form.get('notes')
    return product_name_entered, department_id_entered, amount_entered, notes_entered


@app.route('/add_product', methods=['POST', 'GET'])  # adds a product to list
@login_required
def add_product():
    """
                Add the new product details to the database.If the name of the product exists a compatible message will follow.

                Returns: redirect to the function 'show_products' that shows all products.

                """

    if request.method == "POST":
        product_name_entered, department_id_entered, amount_entered, notes_entered = get_details_product()
        if product_name_entered.isnumeric():
            flash('you can only use letters while adding products')
            return redirect(url_for('show_products'))
        else:
            Qf = QueriesMyDb(Products, product_name_entered, Products.product_name)
            one_record_details = Qf.query_filter_name()
            if one_record_details:  # if exists, give a message.
                flash(f'The product {one_record_details.product_name} exists and was not added to the list !!! ')
                return redirect(url_for('show_products'))
            else:
                product_details_entered = Products(product_name_entered, department_id_entered, amount_entered,
                                                   notes_entered, None)
                db.session.add(product_details_entered)
                db.session.commit()
                return redirect(url_for('show_products'))


@app.route('/add_department', methods=['POST', 'GET'])  # adds a department
@login_required
def add_departments():
    """
               Adds new department category, if department does not exist.

                Returns:
                Returns: redirect to the function 'show_products' that shows all products.

     """

    dict_departments = session['dict_departments']
    if request.method == "POST":
        try:
            department_name_entered = request.form.get("department_name")
            assert re.match(r'[A-Za-z ]+$', department_name_entered)!=None
        except AssertionError:
            flash(f'The department {department_name_entered} you entered must be letters only, try again.')
            return redirect(url_for('show_products'))
        else:
            if department_name_entered in dict_departments.values():
                flash(f'The department {department_name_entered} already exists')
                return redirect(url_for('show_products'))
            else:
                department_to_insert = Departments(len(dict_departments) + 1, department_name_entered)
                db.session.add(department_to_insert)
                db.session.commit()
                flash(f'The department {department_name_entered} was added successfully')
                return redirect(url_for('col_names_dict_names_to_session'))
    return redirect(url_for('col_names_dict_names_to_session'))


@app.route('/add_sub/<int:product_id>/<operation>', methods=['GET'])  # adds amount or subtracts
@login_required
def add_sub(product_id, operation):
    """
               Adds or  subtracts  the amount of the product.can't subtruct a zero amount, a compatible message will appear.
               Parameters:
                  product_id(int): The product id.
                  operation(string) :The operation chosen (Add or Sub).
                Returns:
                Returns: redirect to the function 'show_products' that shows all products.
    """
    if request.method == 'GET':
        product_chosen_id = Products.query.get(product_id)
        if operation == 'Add':
            product_chosen_id.product_amount += 1
            db.session.commit()
            return redirect(url_for('show_products'))
        elif operation == 'Subtruct':
            if product_chosen_id.product_amount == 0:
                flash(f'you can only add the amount for this product !!! ')
            else:
                product_chosen_id.product_amount -= 1
                db.session.commit()
            return redirect(url_for('show_products'))
        else:
            return render_template('index_products.html')
    else:

        return redirect(url_for('show_products'))


@app.route('/del_product/<int:product_id>', methods=['POST', 'GET'])
@login_required
def del_product(product_id):
    """
                 Deletes a product.
                  Parameters:
                  product_id(int): The product id.
                  Returns:
                  Returns: redirect to the function 'show_products' that shows all products.
       """
    if request.method == 'GET':
        pro_to_del_id = Products.query.get(product_id)
        db.session.delete(pro_to_del_id)
        db.session.commit()
        return redirect(url_for('show_products'))
    else:
        return redirect(url_for('show_products'))


@app.route('/dell_all', methods=['POST', 'GET'])
@login_required
def dell_all():
    """
                      Deletes all products.

                      Returns:
                      Returns: redirect to the function 'show_products' that shows all products.
    """
    if request.method == 'GET':
        Products.query.delete()
        db.session.commit()
        return redirect(url_for('show_products'))
    else:
        return redirect(url_for('show_products'))


@app.route('/find_product', methods=['POST', 'GET'])
@login_required
def find_product():
    """
                      Finds a product in the database.If the product is not found  a compatible message will appear.
                      Returns:
                      Returns: redirect to the function 'show_products' that shows all products.
    """
    dict_products_by_department = dict()
    list_column_names = session['list_column_names']
    dict_departments = session['dict_departments']
    if request.method == 'POST':
        product_name_entered = request.form.get('product_name')
        Q1 = QueriesMyDb(Products, product_name_entered, Products.product_name)
        find_product = Q1.query_filter_name()
        if find_product:
            for num_department in dict_departments:
                if num_department == str(find_product.department_id):#if department number equals the find product,department.
                    find_dept = dict_departments[num_department]
                    dict_products_by_department = inserts_products_to_dict(num_department, dict_departments,
                                                                           dict_products_by_department)#finds product tha belong to department of find prod.

            return render_template('index_products.html', list_column_names=list_column_names,
                                   dict_products_by_department=dict_products_by_department
                                   , dict_departments=dict_departments, find_name=find_product.product_name,
                                   find_dept=find_dept)
        else:
            flash(f'The product {product_name_entered} does not exist the list !!! ')
            return redirect(url_for('show_products'))




@app.route('/export_list/<date>', methods=['POST', 'GET'])
@login_required
def export_list(date):
    """                      Exports the list to an excel sheet.List can't be exported if there are no products.
                             parameters:
                             date(date):The date today.
                             Returns:
                             Returns: redirect to the url 'show_products' that shows all products .
    """
    try:
        assert Products.query.all()
    except AssertionError:
        flash("You can not export list.")
        return redirect(url_for('show_products'))
    else:
        list_column_atrb = ["product_id", "product_name", "department_id", "product_amount", "product_notes",
                            "date_buy"]
        dict_products_by_department = dict()
        list_column_names = session['list_column_names']
        dict_departments = session['dict_departments']
        recievef = openpyxl.load_workbook('product_lists.xlsx')
        g_sheets_f = recievef.sheetnames #get the  sheet names that were in the beginning.
        num_sheets_start = len(g_sheets_f)
        recievef.create_sheet('list' + date)
        g_sheets = recievef.sheetnames
        sheet_n = recievef.get_sheet_by_name(g_sheets[-1])#get the last sheet name to update it.
        sheet_n.cell(row=1, column=1).value = " Shopping  list"
        for num_department in dict_departments:
            dict_products_by_department = inserts_products_to_dict(num_department, dict_departments,
                                                                   dict_products_by_department)
        for key_dep_name in dict_products_by_department:  # writing  the name of department to excel file
            sheet_n.cell(row=sheet_n.max_row + 1, column=1).value = key_dep_name
            sheet_n['A1'].font = Font(bold=True)  # emphasize the title.
            sheet_n['A' + str(sheet_n.max_row)].font = Font(bold=True)  # emphasize the department name.
            row = sheet_n.max_row + 1
            for index, column_name in enumerate(
                    list_column_names):  # writing  the column names to excel file.runs every department roumd in first for loop.
                sheet_n.cell(row=row, column=index + 1).value = list_column_names[index]
                colum_letter = string.ascii_uppercase[index]#get the capital for the index at the loop.
                sheet_n[colum_letter + str(row)].font = Font(bold=True)# bold the titles of the detail product.
            for value_product in dict_products_by_department[
                key_dep_name]:  # takes the value_product_object and writing the product details to excel file.
                row += 1
                for colum_number in range(1, len(list_column_names) + 1):
                    sheet_n.cell(row=row, column=colum_number).value = getattr(value_product, list_column_atrb[
                        colum_number - 1])  # using the name of the attributes from  list_column_atrb
        num_rows_excell = sheet_n.max_row
        num_sheets_fin = len(g_sheets)
        recievef.save('product_lists.xlsx')

        return redirect(url_for('check_export', num_rows_excell=num_rows_excell, num_sheets_start=num_sheets_start,
                                num_sheets_fin=num_sheets_fin,
                                length_dict_products_by_department=len(dict_products_by_department)))


# arguments:,number of sheet before export,number of sheet after export,length of dic - to know how many departments implemented.
@app.route(
    '/check_export/<int:num_rows_excell>/<int:num_sheets_fin>/<int:num_sheets_start>/<int:length_dict_products_by_department>',
    methods=['POST', 'GET'])
@login_required
def check_export(num_rows_excell, num_sheets_fin, num_sheets_start, length_dict_products_by_department):
    """                         Checks products list was exported. the user receives a message according if it succeeded or not.
                                 parameters:
                                 num_rows_excell(int):The number of rows active on current excell sheet .
                                 num_sheets_start (int): number of sheet before export.
                                 num_sheets_fin (int) : number of sheets after export.
                                 length_dict_products_by_department(int) : length of dic - to know how many departments implemented
                                 Returns:
                                 Returns: redirect to the url 'show_products' that shows all products .
        """
    num_rows_export = (db.session.query(Products.product_name).count()) + ((
                                                                               length_dict_products_by_department) * 2)  # calculating the number of rows required to export.length dict*2(department+row of titles)
    if (num_sheets_fin) - num_sheets_start == 1 and num_rows_export == (
            num_rows_excell - 1):  # if rows expected equal rows actual and sheet was add.
        flash(f'The list was exported successfully ')
    else:
        flash(f'There was an error exporting the list ')

    return redirect(url_for('show_products'))


@app.route('/import_list', methods=['GET', 'POST'])
@login_required
def import_list():
    """
                              Imports  a product list from excel by choosing the name of the sheet.Only if there are no items.

                              Returns:
                              Returns: redirect to the url 'show_products' that shows all products.

    """
    if Products.query.all():
        flash("You can  not import list. you must delete it first")
        return redirect(url_for('show_products'))
    else:
        if request.method == "POST":
            sheet_name_entered = request.form.get('sheet_select')
            engine = create_engine("mysql://root:newrootpassword@localhost/home_products_management")
            names = ['Id_Product', 'Product_Name', 'Id_Department', 'Product_amount', 'Product_Notes', 'Date_Buy']
            excel_file = 'C:/Users/anatei/PycharmProjects/HelloWorld/FLASK_APP/product_lists.xlsx'
            product_sheet = pd.read_excel(excel_file, sheet_name=sheet_name_entered, index_col=0, header=None,
                                          names=names)
            product_sheet_s = product_sheet.sort_values(by=['Product_Name'], ascending=False)#sort by the column.gives all titles under the database.
            len_excel_titeles = (len(product_sheet_s.loc['Id_Product'])) * 2 + 1# get all lines which contain Id_Product.muls by two in order to calc the dept title as well. and add the sheet main title.
            product_sheet_s.iloc[:-len_excel_titeles].to_sql(con=engine, name="products", if_exists='append',
                                                             index=True)#from the start get the data till you reach all the titles.
            return redirect(url_for('show_products'))
        return redirect(url_for('show_products'))


@app.route('/date_buy', methods=['Post', 'GET'])
@login_required
def date_buy():
    """
                                     Inserts the date today into the date column

                                     Returns: redirect to the url 'show_products' that shows all products .
     """
    all_products_data = Products.query.all()
    for data in all_products_data:
        data.date_buy = date.today()
        db.session.commit()
    return redirect(url_for('col_names_dict_names_to_session'))


@app.route('/show_new_recipes',methods=['POST','GET'])
@login_required
def show_new_recipes():
    """
                          Shows all the recipes and their products.
                          Returns:
                          Returns: render_template to the page that shows all recipes.
    """
    if request.method =='GET':
       list_column_names_prod_recipes = session['list_column_names_prod_recipes']
       dict_departments = session['dict_departments']
       all_recipes = db.session.query(Recipes).join(ProdForRecipe, ProdForRecipe.recipe_id == Recipes.recipe_id).all()
       all_products_for_recipes = ProdForRecipe.query.all()
       dict_recipes = {rec.recipe_id: rec.recipe_name for rec in Recipes.query.all()}
       return render_template('show_recipes.html',all_recipes=all_recipes,all_products_for_recipes= all_products_for_recipes,list_column_names_prod_recipes=list_column_names_prod_recipes,dict_departments=dict_departments,dict_recipes=dict_recipes)
    return redirect(url_for('col_names_dict_names_to_session'))


@app.route('/add_new_recipe_to_list/<recipe_id>', methods=['POST', 'GET'])  # adds products of recipe chosen to list.
@login_required
def add_new_recipe_to_list(recipe_id):
    prod_found = ""
    if request.method=='GET':
       all_products_recipe = ProdForRecipe.query.filter_by(recipe_id=recipe_id).all()
       for product in all_products_recipe:

           query_product = QueriesMyDb(Products,product.product_name, Products.product_name)
           if query_product.query_filter_name():
              print(product.product_name)
              prod_found += (product.product_name + " ,")  # if it exists than get a message into variable prod_found
              print(prod_found)
           else:
             products_from_recipe = Products(product.product_name,product.department_id_rec,product.product_amount,product.product_notes,None)
             db.session.add(products_from_recipe)
             db.session.commit()
       if prod_found:
           flash(f'The products: {prod_found} exists and were not added to the list !!! ')
       return redirect(url_for('col_names_dict_names_to_session'))

    return redirect(url_for('col_names_dict_names_to_session'))


@app.route('/add_new_prod_recipe/get_detail', methods=['POST', 'GET'])  # adds products of recipe chosen to list.
@login_required
def get_details_product_recipe():
       product_name_entered = request.form.get('product_name')
       department_name_entered = request.form.get('department_select')  # user selects department name
       dict_departments = session['dict_departments']
       for key_id in dict_departments:  # translating department name chosen to department id.
            if department_name_entered == dict_departments[key_id]:  # if finds the name of department in dict.get its number.
                department_id_entered = key_id
       amount_entered = request.form.get('amount')
       notes_entered = request.form.get('notes')
       recipe_entered = request.form.get('recipe_select')
       record_recipe = Recipes.query.filter_by(recipe_name = recipe_entered).first()
       num_rec_id  = record_recipe.recipe_id
       return product_name_entered, department_id_entered, amount_entered,notes_entered,recipe_entered,num_rec_id



@app.route('/add_new_prod_recipe', methods=['POST', 'GET'])  # adds products of recipe chosen to list.
@login_required
def add_new_prod_recipe():
    if request.method == 'POST':
       product_name_entered, department_id_entered, amount_entered, notes_entered, recipe_entered,num_rec_id = get_details_product_recipe()
       if product_name_entered.isnumeric():
          flash('you can only use letters while adding products')
          return redirect(url_for('show_products'))
       else:
           Qr = QueriesMyDb(ProdForRecipe,product_name_entered, ProdForRecipe.product_name)
           one_record_details = Qr.query_filter_name()
           if one_record_details:  # if exists, give a message.
               flash(f'The product {one_record_details.product_name} exists and was not added to the list !!! ')
               return redirect(url_for('show_products'))
           else:
               product_details_entered = ProdForRecipe(product_name_entered, department_id_entered, amount_entered,notes_entered,num_rec_id)
               db.session.add(product_details_entered)
               db.session.commit()
               return redirect(url_for('show_new_recipes'))

    return redirect(url_for('show_new_recipes'))



@app.route('/add_new_recipe',methods=['POST','GET'])
def add_new_recipe():
    dict_recipes = session['dict_recipes']
    if request.method=='POST':
        try:
            new_rec_entered = request.form.get('recipe_name')
            assert re.match(r'[A-Za-z ]+$', new_rec_entered)!=None
        except AssertionError:
            flash(f'The Recipe {new_rec_entered} you entered must be letters only, try again.')
            return redirect(url_for('show_products'))
        else:
            if new_rec_entered in dict_recipes.values():
               flash(f'The recipe {new_rec_entered} already exists')
               return redirect(url_for('show_new_recipes'))
            else:
                results = [item.recipe_id for item in Recipes.query.order_by(desc(Recipes.recipe_id)).all()]
                recipe_to_insert = Recipes(new_rec_entered,int(results[0]) + 1)
                db.session.add(recipe_to_insert)
                db.session.commit()
                flash(f'The recipe {new_rec_entered} was added successfully')
                return redirect(url_for('show_new_recipes'))
    return redirect(url_for('show_new_recipes'))


@app.route('/show_my_recipes',methods=['POST','GET'])
def show_my_recipes():
       dict_recipes = {rec.recipe_id: rec.recipe_name for rec in Recipes.query.all()}
       list_column_names_recipes = session['list_column_names_recipes']
       return render_template('show_my_recipes.html',dict_recipes=dict_recipes,list_column_names_recipes=list_column_names_recipes)


@app.route('/del_rec/<int:recipe_id>', methods=['POST', 'GET'])
@login_required
def del_rec(recipe_id):
    """
                 Deletes a recipe.
                  Parameters:
                  recipe_id(int): The recipe id.
                  Returns: redirect to the function 'show_new_recipes' that shows all recipes products.
       """
    if request.method == 'GET':
       Qd = QueriesMyDb(ProdForRecipe, recipe_id, ProdForRecipe.recipe_id)
       one_record_details = Qd.query_filter_name()
       try:
           assert one_record_details!=None
       except:
           rec_to_del_id = Recipes.query.get(recipe_id)
           db.session.delete(rec_to_del_id)
           db.session.commit()
           return redirect(url_for('show_new_recipes'))

       else:
          products_under_recipe = ProdForRecipe.query.filter_by(recipe_id = recipe_id).all()
          for prod_rec in products_under_recipe:
             db.session.delete(prod_rec)
          rec_to_del_id = Recipes.query.get(recipe_id)
          db.session.delete(rec_to_del_id)
          db.session.commit()
          return redirect(url_for('show_new_recipes'))

    return redirect(url_for('show_new_recipes'))

